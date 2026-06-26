"""
Match-kvalitetsrapport for BeerSniffer.

Scanner data.json og finder potentielle matchproblemer:
  1. OPSPLITTET   — samme øl (navn+volumen+ABV) ligger i flere grupper hos
                    forskellige butikker. Burde måske være slået sammen.
  2. INTERN ROD   — en gruppe indeholder priser med forskellig volumen eller
                    vidt forskellig pris-pr-liter (= sandsynlig fejlmatch).
  3. IKKE-ØL      — navne der ligner merchandise/bundles der slap igennem.

Skriver match_rapport.xlsx (samme stil som fejlliste.xlsx). Read-only:
ændrer intet i data.json — det er en tjekliste til dig.

Kør:  python match_rapport.py
"""

import json
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz
    HAR_FUZZ = True
except ImportError:
    HAR_FUZZ = False

DATA = "data.json"
OUTPUT = "match_rapport.xlsx"

# ---- Farver (samme palet som fejlliste) ----
HEAD_FILL = PatternFill("solid", fgColor="2E5E4E")
HIGH_FILL = PatternFill("solid", fgColor="F4A6A6")   # rød = alvorligt
MED_FILL = PatternFill("solid", fgColor="FFF3B0")    # gul = tjek
LOW_FILL = PatternFill("solid", fgColor="E3F0E3")    # grøn = info
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)

COLUMNS = [
    ("Type", 16),
    ("Alvor", 10),
    ("Beskrivelse", 40),
    ("Øl A", 38),
    ("Butik(ker) A", 22),
    ("URL A", 45),
    ("Øl B", 38),
    ("Butik(ker) B", 22),
    ("URL B", 45),
    ("Detaljer", 30),
]


def _norm(s):
    s = (s or "").lower()
    s = re.sub(r'\d+(?:[.,]\d+)?\s*(cl|ml|l|%)', ' ', s)
    s = re.sub(r'[^a-zæøå0-9 ]', ' ', s)
    for w in ['trappist', 'brouwerij', 'der', 'van', 'co', 'brewing',
              'brewery', 'the', 'bryggeri']:
        s = re.sub(rf'\b{w}\b', ' ', s)
    return ' '.join(s.split())


def _shops(beer):
    return ", ".join(p['shop_name'] for p in beer.get('prices', []))


def _urls(beer):
    """Saml produkt-URL'er fra en gruppes priser (én pr. linje)."""
    urls = [p.get('url', '') for p in beer.get('prices', []) if p.get('url')]
    return "\n".join(urls)


def _navnescore(a, b):
    if not HAR_FUZZ:
        # simpel fallback: andel fælles ord
        sa, sb = set(_norm(a).split()), set(_norm(b).split())
        if not sa or not sb:
            return 0
        return round(100 * len(sa & sb) / max(len(sa), len(sb)))
    return max(fuzz.token_sort_ratio(_norm(a), _norm(b)),
               fuzz.token_set_ratio(_norm(a), _norm(b)))


def find_opsplittede(beers):
    """Samme kerneord+volumen+afrundet ABV, men separate grupper, flere butikker."""
    buckets = defaultdict(list)
    for b in beers:
        words = tuple(sorted(set(_norm(b.get('name', '')).split())))
        if not words:
            continue
        key = (words, b.get('volume_cl'), round(b.get('abv') or 0))
        buckets[key].append(b)

    fund = []
    for key, gruppe in buckets.items():
        if len(gruppe) < 2:
            continue
        butikker = set()
        for g in gruppe:
            butikker.update(p['shop_name'] for p in g.get('prices', []))
        if len(butikker) > 1:
            # rapportér parvis
            for i in range(len(gruppe)):
                for j in range(i + 1, len(gruppe)):
                    a, c = gruppe[i], gruppe[j]
                    sc = _navnescore(a.get('name', ''), c.get('name', ''))
                    fund.append({
                        "Type": "OPSPLITTET",
                        "Alvor": "HØJ" if sc >= 85 else "MELLEM",
                        "Beskrivelse": "Samme øl ligger i 2 grupper",
                        "Øl A": a.get('name', ''),
                        "Butik(ker) A": _shops(a),
                        "URL A": _urls(a),
                        "Øl B": c.get('name', ''),
                        "Butik(ker) B": _shops(c),
                        "URL B": _urls(c),
                        "Detaljer": f"navne-lighed {sc} | {key[1]}cl | ~{key[2]}%",
                    })
    return fund


def find_intern_rod(beers):
    """Grupper hvor priserne har forskellig volumen eller mistænkelig pris-spredning."""
    fund = []
    for b in beers:
        prices = b.get('prices', [])
        if len(prices) < 2:
            continue
        # pris-spredning: største vs mindste
        pp = [p['price'] for p in prices if p.get('price')]
        if pp and min(pp) > 0:
            spread = max(pp) / min(pp)
            if spread >= 2.0:   # mere end dobbelt så dyrt = mistænkeligt
                fund.append({
                    "Type": "PRIS-SPREDNING",
                    "Alvor": "MELLEM",
                    "Beskrivelse": "Stor prisforskel i samme gruppe",
                    "Øl A": b.get('name', ''),
                    "Butik(ker) A": _shops(b),
                    "URL A": _urls(b),
                    "Øl B": "",
                    "Butik(ker) B": "",
                    "URL B": "",
                    "Detaljer": f"{min(pp):.0f}–{max(pp):.0f} kr ({spread:.1f}x)",
                })
    return fund


def find_ikke_oel(beers):
    """Navne der ligner merchandise/bundles MEN ikke allerede er håndteret.

    Springer over produkter der allerede er korrekt kategoriseret som
    'smagekasse' — de er ikke et problem, de vises bare på smagekasse-fanen.
    Matcher signalord som HELE ord (så 'krus' ikke rammer 'Krush').
    """
    signaler = ["glas", "krus", "drikkehorn", "riedel", "abonnement",
                "gavekort", "merchandise", "forsendelse", "fragt", "t-shirt"]
    fund = []
    for b in beers:
        # Allerede korrekt markeret som smagekasse? Så er det ikke et problem.
        if b.get('category') == 'smagekasse':
            continue
        n = (b.get('name') or '').lower()
        ramt = [s for s in signaler if re.search(rf'\b{re.escape(s)}\b', n)]
        # Smagekasse/bundle-ord der IKKE er fanget af kategorien => fejl
        for kasse_ord in ["smagekasse", "smagkasse", "bland selv"]:
            if kasse_ord in n:
                ramt.append(kasse_ord)
        if ramt:
            fund.append({
                "Type": "IKKE-ØL?",
                "Alvor": "HØJ",
                "Beskrivelse": "Ligner ikke-øl der slap igennem",
                "Øl A": b.get('name', ''),
                "Butik(ker) A": _shops(b),
                "URL A": _urls(b),
                "Øl B": "",
                "Butik(ker) B": "",
                "URL B": "",
                "Detaljer": f"signalord: {', '.join(ramt)}",
            })
    return fund


def main():
    data = json.load(open(DATA, encoding='utf-8'))
    beers = data['beers']

    rows = []
    rows += find_ikke_oel(beers)
    rows += find_opsplittede(beers)
    rows += find_intern_rod(beers)

    # Sortér: alvor (HØJ først), så type
    alvor_rank = {"HØJ": 0, "MELLEM": 1, "LAV": 2}
    rows.sort(key=lambda r: (alvor_rank.get(r["Alvor"], 9), r["Type"]))

    _skriv_xlsx(rows, beers)

    print(f"✅ Scannede {len(beers)} grupper")
    print(f"   Opsplittede par:  {sum(1 for r in rows if r['Type']=='OPSPLITTET')}")
    print(f"   Pris-spredning:   {sum(1 for r in rows if r['Type']=='PRIS-SPREDNING')}")
    print(f"   Ikke-øl?:         {sum(1 for r in rows if r['Type']=='IKKE-ØL?')}")
    print(f"   I alt:            {len(rows)} → {OUTPUT}")


def _skriv_xlsx(rows, beers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Match-rapport"

    for ci, (head, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=head)
        c.fill = HEAD_FILL
        c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    keys = [h for h, _ in COLUMNS]
    for ri, rec in enumerate(rows, start=2):
        fill = {"HØJ": HIGH_FILL, "MELLEM": MED_FILL, "LAV": LOW_FILL}.get(rec["Alvor"], None)
        for ci, key in enumerate(keys, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(key, ""))
            c.font = BODY
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(key in ("Øl A", "Øl B", "Beskrivelse", "Detaljer", "URL A", "URL B")))
            if ci == 2 and fill:   # farv kun Alvor-cellen
                c.fill = fill

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows)+1}"
    ws.row_dimensions[1].height = 28

    # Oversigts-fane
    ws2 = wb.create_sheet("Oversigt")
    ws2.column_dimensions["A"].width = 90
    intro = [
        ("Match-rapport — sådan læser du den", True),
        ("", False),
        ("Rapporten finder potentielle problemer i grupperingen. Den ændrer", False),
        ("INTET i data.json — det er en tjekliste.", False),
        ("", False),
        ("OPSPLITTET = samme øl ligger i flere grupper hos forskellige", False),
        ("  butikker. HØJ alvor = navnene ligner hinanden meget (burde nok", False),
        ("  samles). Skyldes typisk rod i bryggeri-feltet.", False),
        ("", False),
        ("PRIS-SPREDNING = priserne i samme gruppe varierer over 2x. Kan", False),
        ("  være en fejlmatch (to forskellige øl samlet), eller bare et godt", False),
        ("  tilbud. Tjek manuelt.", False),
        ("", False),
        ("IKKE-ØL? = navnet indeholder ord som 'glas', 'bland selv',", False),
        ("  'smagekasse' osv. Bør sandsynligvis skippes i scraperen.", False),
        ("", False),
        ("Farve i Alvor-kolonnen: rød = kig på den, gul = tjek hvis du har tid.", False),
    ]
    for i, (t, b) in enumerate(intro, start=1):
        c = ws2.cell(row=i, column=1, value=t)
        c.font = Font(name="Arial", bold=b, size=12 if b else 10,
                      color="2E5E4E" if b else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT)


if __name__ == "__main__":
    main()