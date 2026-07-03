# -*- coding: utf-8 -*-
"""
flag_bryggeri_stoej.py

FLAGGER (retter ikke) forurenede Bryggeri-vaerdier i fejlliste.xlsx,
saa du selv kan rette dem i ro og mag.

Hvorfor Noter-kolonnen?
  write_fejlliste() genskriver arket ved hvert build, og Bryggeri-cellen
  overskrives med scraperens (forurenede) vaerdi. Noter er det ENESTE
  redigerbare felt der bevares paa tvaers af builds — saa flaget der
  bliver staaende, indtil du har rettet.

Hvad flagges (samme heuristik som diagnosen):
  - Butiksnavn i Bryggeri (fx 'Beershoppen' paa en De Dolle-oel)
  - Produktnavns-stoej: komma+aarstal/stilord, volumen/ABV-rester,
    afklippede vaerdier der ender paa komma/streg

Saadan retter du bagefter:
  1. Aabn fejlliste.xlsx, filtrer Noter-kolonnen paa '⚠ RET BRYGGERI'
     (autofilter er allerede slaaet til).
  2. Skriv det KORREKTE bryggeri i Bryggeri-kolonnen (gylden overskrift).
     Din vaerdi bliver facit og vinder over scraperen for evigt.
  3. Slet evt. flaget fra Noter naar rettet (valgfrit — det er kun tekst).
  4. python build_data.py

Sikkerhed:
  - --dry-run foerst: viser alle raekker der ville blive flagget
  - .bak-kopi foer skrivning (fejlliste.xlsx.bak_flag)
  - Eksisterende noter OVERSKRIVES ALDRIG — flaget saettes foran
  - Raekker der allerede baerer flaget springes over (koer den gerne igen)
  - Luk Excel foer koersel med skrivning

Koer fra roden (SEPARATE linjer):
    python flag_bryggeri_stoej.py --dry-run
    python flag_bryggeri_stoej.py
"""

import re
import shutil
import sys
from html import unescape

from openpyxl import load_workbook

XLSX_FILE = "fejlliste.xlsx"
XLSX_BACKUP = "fejlliste.xlsx.bak_flag"

FLAG = "\u26a0 RET BRYGGERI"

URL_HEADER_PREFIX = "URL"
BREW_HEADER = "Bryggeri"
SHOP_HEADER = "Butik"
NAME_HEADER = "Navn"
NOTER_HEADER = "Noter"

SHOP_NAMES = {
    "brygshoppen", "beermatch", "drikbeer", "a good case", "agoodcase",
    "beershoppen", "best of beers", "bestofbeers", "oeltanken", "øltanken",
    "beer me", "beerme", "vild med vin", "vildmedvin",
}

YEAR_RE = re.compile(r"\b20(1[5-9]|2[0-9])\b")
STYLE_WORDS = re.compile(
    r"\b(ale|stout|porter|ipa|neipa|lager|pilsner|pils|tripel|dubbel|quad|"
    r"quadrupel|saison|sour|gueuze|geuze|lambic|lambik|bock|weissbier|"
    r"weizen|blond|blonde|kriek|cider|maltoel|trappist)\b",
    re.IGNORECASE,
)
VOLABV_RE = re.compile(
    r"\d+\s*(?:x\s*\d+\s*)?(cl|ml|l)\b\.?|\d+[.,]?\d*\s*%", re.IGNORECASE
)


def is_polluted(brew: str, shop: str = "") -> tuple[bool, str]:
    """(forurenet?, aarsag)"""
    if not brew:
        return False, ""
    b = unescape(brew).strip()
    if b.lower() in SHOP_NAMES:
        return True, "butiksnavn"
    if shop and b.lower() == shop.strip().lower():
        return True, "butiksnavn"
    if "," in b and (YEAR_RE.search(b) or STYLE_WORDS.search(b)):
        return True, "produktnavn i feltet"
    if b.endswith((",", "-", "\u2013", "\u2014")):
        return True, "afklippet vaerdi"
    if VOLABV_RE.search(b):
        return True, "volumen/ABV-rest"
    return False, ""


def main():
    dry_run = "--dry-run" in sys.argv
    if dry_run:
        print("🔍 DRY RUN — fejlliste.xlsx roeres ikke.\n")

    try:
        wb = load_workbook(XLSX_FILE)
    except FileNotFoundError:
        print(f"❌ Fandt ikke {XLSX_FILE} — koer fra projektets rod.")
        sys.exit(1)
    except PermissionError:
        print(f"❌ {XLSX_FILE} er laast — luk Excel og proev igen.")
        sys.exit(1)

    ws = wb.active
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1) if c.value}
    brew_col = headers.get(BREW_HEADER)
    shop_col = headers.get(SHOP_HEADER)
    name_col = headers.get(NAME_HEADER)
    noter_col = headers.get(NOTER_HEADER)
    if not all([brew_col, shop_col, name_col, noter_col]):
        print(f"❌ Fandt ikke alle kolonner. Headers: {list(headers)}")
        sys.exit(1)

    flagged, already = [], 0
    for row in ws.iter_rows(min_row=2):
        name = row[name_col - 1].value
        if not name:
            continue
        brew = str(row[brew_col - 1].value or "").strip()
        shop = str(row[shop_col - 1].value or "").strip()
        noter_cell = row[noter_col - 1]
        noter = str(noter_cell.value or "").strip()

        polluted, reason = is_polluted(brew, shop)
        if not polluted:
            continue
        if FLAG in noter:
            already += 1
            continue

        flag_text = f"{FLAG} ({reason}): {brew!r}"
        new_noter = f"{flag_text} | {noter}" if noter else flag_text
        flagged.append((str(name), brew, reason))
        if not dry_run:
            noter_cell.value = new_noter

    print(f"🚩 Raekker der {'ville blive' if dry_run else 'blev'} flagget i Noter: "
          f"{len(flagged)}")
    if already:
        print(f"   (sprunget over — allerede flagget: {already})")
    print()
    for name, brew, reason in flagged:
        print(f"   {unescape(name)[:68]}")
        print(f"      Bryggeri nu: {brew!r}   [{reason}]")

    if dry_run:
        print("\nSer listen rigtig ud, koer:  python flag_bryggeri_stoej.py")
        return

    shutil.copy2(XLSX_FILE, XLSX_BACKUP)
    print(f"\n💾 Backup gemt: {XLSX_BACKUP}")
    wb.save(XLSX_FILE)
    print(f"✅ {XLSX_FILE} opdateret — {len(flagged)} raekker flagget.")
    print(f"""
Saadan retter du:
  1. Aabn arket, filtrer Noter paa '{FLAG}'
  2. Skriv korrekt bryggeri i Bryggeri-kolonnen (bliver permanent facit)
  3. python build_data.py naar du har rettet en portion
Flaget i Noter kan du selv slette efterhaanden — eller lade staa.
Rul evt. tilbage: copy {XLSX_BACKUP} {XLSX_FILE}""")


if __name__ == "__main__":
    main()