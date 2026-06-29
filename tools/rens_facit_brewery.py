"""
Renser fejlliste.xlsx: nulstiller brewery-celler der ved en fejl indeholder
holdbarhedsdatoer ('Best Before: ...') i stedet for et bryggerinavn.

- Laver en backup (fejlliste_foer_rens.xlsx) FOER noget aendres.
- Roerer KUN Bryggeri-celler der matcher dato/best-before-moenstret.
- Volumen, ABV, noter og korrekte bryggerier roeres ikke.
- Spoerger 'skriv JA' foer den gemmer.

Koeres fra repo-roden:  python tools\rens_facit_brewery.py
"""
import os
import re
import shutil
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "fejlliste.xlsx")

# Matcher: 'best before', 'holdbar', datoer som 08/01-2027 eller 06-08-26, 'udloeber'
BAD = re.compile(
    r"best\s*before|holdbar|\d{2}[/-]\d{2}[/-]\d{2,4}|udl[\u00f8o]ber",
    re.I,
)


def main():
    if not os.path.exists(XLSX):
        print(f"FEJL: finder ikke {XLSX}")
        return

    backup = os.path.join(ROOT, "fejlliste_foer_rens.xlsx")
    shutil.copy(XLSX, backup)
    print(f"Backup gemt: {backup}")

    wb = load_workbook(XLSX)
    ws = wb["Fejlliste"] if "Fejlliste" in wb.sheetnames else wb.active

    headers = [c.value for c in ws[1]]
    if "Bryggeri" not in headers:
        print("FEJL: fandt ikke 'Bryggeri'-kolonnen i arket.")
        print(f"Kolonner: {headers}")
        return
    bcol = headers.index("Bryggeri") + 1

    # find ogsaa Navn-kolonnen til paen visning (valgfrit)
    ncol = headers.index("Navn") + 1 if "Navn" in headers else None

    to_clear = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=bcol).value
        if v is not None and BAD.search(str(v)):
            name = ws.cell(row=r, column=ncol).value if ncol else ""
            to_clear.append((r, v, name))

    print(f"\nFandt {len(to_clear)} brewery-celler der ser ud som datoer.\n")
    print("Foerste 12:")
    for r, v, name in to_clear[:12]:
        navn = (str(name)[:45]) if name else ""
        print(f"   raekke {r:>4}: [{v}]   <- {navn}")

    if not to_clear:
        print("\nIntet at rense. Arket er uaendret.")
        return

    print()
    svar = input(f"Nulstil disse {len(to_clear)} brewery-celler? (skriv JA): ").strip()
    if svar != "JA":
        print("Afbrudt. Arket er IKKE aendret.")
        return

    for r, v, name in to_clear:
        ws.cell(row=r, column=bcol).value = None

    wb.save(XLSX)
    print(f"\nFaerdig: {len(to_clear)} celler nulstillet.")
    print(f"Backup ligger i: {os.path.basename(backup)}")
    print("De nulstillede oel dukker op som 'MANGLER bryggeri' naeste gang - korrekt.")


if __name__ == "__main__":
    main()