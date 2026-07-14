"""
Facitliste-/override-system til BeerSniffer.

Én fil styrer det hele: fejlliste.xlsx
  - Du udfylder de manglende felter (volumen, bryggeri, pak-antal, abv).
  - Næste scraping læser arket og bruger dine værdier som facit.
  - Kun NYE øl med manglende felter dukker op som "MANGLER". Dine rettelser
    huskes, fordi nøglen er butik+URL (stabil selv om navnet ændrer sig).

Flow i export_json.py:
  existing = load_fejlliste("fejlliste.xlsx")
  for item in items:
      apply_overrides(item, existing)   # FØR matchning
  ... match/gruppér ...
  write_fejlliste(items, existing, "fejlliste.xlsx")
"""

import os
import re
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Felter brugeren kan udfylde (override vinder altid over scraper) ----
OVERRIDE_FIELDS = ["volume_cl", "pack_count", "brewery", "abv"]

# Felter der tæller som "MANGLER" (blokerer god matchning). abv er valgfri.
REQUIRED_FOR_MATCH = ["volume_cl", "brewery"]  # pak_antal håndteres betinget

# Kolonneopsætning: (overskrift, intern_nøgle, redigerbar?, bredde)
COLUMNS = [
    ("Status",     "_status",     False, 12),
    ("Mangler",    "_mangler",    False, 22),
    ("Butik",      "shop_name",   False, 16),
    ("Navn",       "name",        False, 46),
    ("Volume_cl",  "volume_cl",   True,  11),
    ("Forslag",    "_forslag",    False, 16),
    ("Pak_antal",  "pack_count",  True,  11),
    ("Bryggeri",   "brewery",     True,  22),
    ("ABV",        "abv",         True,  8),
    ("Noter",      "noter",       True,  28),
    ("Ignorer",    "_ignorer",    True,  9),
    ("Sidst_set",  "_sidst_set",  False, 12),
    ("URL (nøgle — rør ikke)", "_key", False, 50),
]

# ---- Farver ----
FILL_HEADER = PatternFill("solid", fgColor="2E5E4E")   # mørkegrøn (auto-kolonner)
EDIT_HEADER = PatternFill("solid", fgColor="C49A2E")   # gylden (udfyld-kolonner)
MISSING_CELL = PatternFill("solid", fgColor="FFF3B0")  # lys gul = mangler
OK_CELL = PatternFill("solid", fgColor="E3F0E3")       # lys grøn
STALE_ROW = PatternFill("solid", fgColor="F0F0F0")     # grå = ikke set i denne kørsel
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF")
DARK_BOLD = Font(name="Arial", bold=True, color="3A2E00")
BODY = Font(name="Arial", size=10)


def item_key(item):
    """Stabil nøgle: URL hvis muligt, ellers butik+navn (mindre stabil)."""
    url = (item.get("url") or "").strip()
    if url:
        return url
    return f"{item.get('shop_name','?')}::{item.get('name','?')}"


MULTIPACK_PATTERNS = [
    r'\d+\s*stk',              # '6 stk', '24 stk'
    r'\d+\s*-?\s*pak',         # '6-pak', '6 pak'
    r'\d+\s*-?\s*pack',        # '6-pack'
    r'\(\s*\d+\s*stk',         # '(6 stk.)'
    r'kasse',
    r'sixpack', r'six\s*pack',
    r'\d+\s*x\s*\d+\s*(cl|ml)',  # '6 x 33 cl' (antal x volumen)
]


def looks_like_multipack(name):
    """
    True kun ved ægte mængdesignaler. Ignorerer collab-'x' (fx
    'Azvex x Dark Element'), som i craft beer betyder samarbejde
    mellem bryggerier, ikke et antal.
    """
    if not name:
        return False
    n = name.lower()
    return any(re.search(p, n) for p in MULTIPACK_PATTERNS)


def _clean(v):
    """Tom celle -> None. Bevar 0 som gyldig værdi."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")  # tillad dansk decimalkomma
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v):
    f = _to_float(v)
    return int(f) if f is not None else None


def _parse_ignorer(v):
    """x / ja / 1 / true -> True. Alt andet -> False."""
    if v is None:
        return False
    return str(v).strip().lower() in ("x", "ja", "1", "true", "y", "yes")


def load_fejlliste(path):
    """Læs eksisterende ark -> {key: row_dict}. Tom dict hvis filen ikke findes."""
    existing = {}
    if not os.path.exists(path):
        return existing
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    key_to_internal = {h: k for (h, k, _e, _w) in COLUMNS}
    idx = {key_to_internal.get(h): i for i, h in enumerate(headers) if h in key_to_internal}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        key = _clean(row[idx["_key"]]) if "_key" in idx else None
        if not key:
            continue
        rec = {
            "volume_cl": _to_float(row[idx["volume_cl"]]) if "volume_cl" in idx else None,
            "pack_count": _to_int(row[idx["pack_count"]]) if "pack_count" in idx else None,
            "brewery": _clean(row[idx["brewery"]]) if "brewery" in idx else None,
            "abv": _to_float(row[idx["abv"]]) if "abv" in idx else None,
            "noter": _clean(row[idx["noter"]]) if "noter" in idx else None,
            "shop_name": _clean(row[idx["shop_name"]]) if "shop_name" in idx else None,
            "name": _clean(row[idx["name"]]) if "name" in idx else None,
            "_ignorer": _parse_ignorer(row[idx["_ignorer"]]) if "_ignorer" in idx else False,
        }
        existing[str(key)] = rec
    return existing


def apply_overrides(item, existing):
    """Udfyld item's felter fra facitlisten. En udfyldt override VINDER."""
    rec = existing.get(item_key(item))
    if not rec:
        return
    for f in OVERRIDE_FIELDS:
        val = rec.get(f)
        if val is not None:
            item[f] = val


def _is_smagekasse(item):
    """Bland-selv-kasser/smagekasser skal ikke flagges — de har hverken
    én størrelse, ét bryggeri eller én ABV."""
    if item.get("category") == "smagekasse":
        return True
    n = (item.get("name") or "").lower()
    return any(s in n for s in [
        "smagekasse", "smagesæt", "smagskasse", "smagspakke",
        "mix smagekasse", "blandet", "bland selv",
        "ølpakke", "bundle", "pakken", "abonnement", "abonnoment",
        "discovery box", "firmaaftaler", "forsendelse",
        "drikkehorn", "(bog)",
        "kasse", "gift box", "luxury box", "beer box", "gavekasse",
        "julekalender", "advent calender", "advent calendar", "adventskalender",
    ])


def _missing_fields(item):
    """Hvilke felter mangler (efter overrides er anvendt)?"""
    # Smagekasser/bundles flagges aldrig — de er ikke enkeltøl
    if _is_smagekasse(item):
        return []
    miss = []
    if _clean(item.get("volume_cl")) is None:
        miss.append("volumen")
    if _clean(item.get("brewery")) is None:
        miss.append("bryggeri")
    # ABV indgår i søgning/filter på sitet -> behandl som påkrævet
    if _clean(item.get("abv")) is None:
        miss.append("abv")
    # pak-antal flages IKKE laengere (scrapere saetter det ikke - override bevares dog)
    return miss


def _volume_hint(item):
    """
    Kvalificeret bud på volumen ud fra fyldevægt (grams), KUN når volumen
    mangler. Returnerer en tekst-hint, aldrig en automatisk værdi.
    grams >= 950 håndteres allerede automatisk i scraperen (75cl), så her
    dækker vi det usikre interval hvor du selv skal bekræfte.
    """
    if _clean(item.get("volume_cl")) is not None:
        return ""
    grams = item.get("grams")
    if not grams:
        return ""
    try:
        g = float(grams)
    except (TypeError, ValueError):
        return ""
    if g <= 450:
        return f"~{int(g)}g → 33cl?"
    if g <= 650:
        return f"~{int(g)}g → 33/44cl?"
    if g <= 850:
        return f"~{int(g)}g → 44/50cl?"
    if g < 950:
        return f"~{int(g)}g → 50cl?"
    return f"~{int(g)}g → 75cl?"


import re as _flag_re

_FLAG_MARKER = "\u26a0 RET BRYGGERI"
_FLAG_SHOP_NAMES = {
    "brygshoppen", "beermatch", "drikbeer", "a good case", "agoodcase",
    "beershoppen", "best of beers", "bestofbeers", "oeltanken", "oltanken",
    "\u00f8ltanken", "beer me", "beerme", "vild med vin", "vildmedvin",
}
_FLAG_YEAR = _flag_re.compile(r"\b20(1[5-9]|2[0-9])\b")
_FLAG_STYLE = _flag_re.compile(
    r"\b(ale|stout|porter|ipa|neipa|lager|pilsner|pils|tripel|dubbel|quad|"
    r"quadrupel|saison|sour|gueuze|geuze|lambic|lambik|bock|weissbier|"
    r"weizen|blond|blonde|kriek|cider|trappist)\b", _flag_re.IGNORECASE)
_FLAG_VOLABV = _flag_re.compile(
    r"\d+\s*(?:x\s*\d+\s*)?(cl|ml|l)\b\.?|\d+[.,]?\d*\s*%", _flag_re.IGNORECASE)


def _is_polluted_brewery(brew, shop=""):
    """True hvis bryggeri-vaerdien er butiksnavn eller produktnavns-stoej."""
    if not brew:
        return False
    b = str(brew).strip()
    if b.lower() in _FLAG_SHOP_NAMES:
        return True
    if shop and b.lower() == str(shop).strip().lower():
        return True
    if "," in b and (_FLAG_YEAR.search(b) or _FLAG_STYLE.search(b)):
        return True
    if b.endswith((",", "-", "\u2013", "\u2014")):
        return True
    if _FLAG_VOLABV.search(b):
        return True
    return False


def _sync_brewery_flag(noter, brewery, shop=""):
    """Fjern '\u26a0 RET BRYGGERI ...'-flaget hvis bryggeriet er rent;
    tilfoej det hvis forurenet. Bevarer alle oevrige noter."""
    noter = (noter or "").strip()
    # fjern et evt. eksisterende flag-segment (op til | eller linjeslut)
    cleaned = _flag_re.sub(
        r"\s*" + _flag_re.escape(_FLAG_MARKER) + r"[^|]*\|?\s*", "", noter
    ).strip(" |").strip()
    if _is_polluted_brewery(brewery, shop):
        flag = f"{_FLAG_MARKER} (butiksnavn/stoej): {str(brewery).strip()!r}"
        return f"{flag} | {cleaned}" if cleaned else flag
    return cleaned or None


def write_fejlliste(items, existing, path):
    """Skriv/opdater fejlliste.xlsx. Bevarer brugerens rettelser og noter."""
    today = date.today().strftime("%d-%m-%Y")

    # Saml rækker: alle øl set i denne kørsel + bevarede rækker (ikke set nu)
    rows = {}
    seen = set()
    for item in items:
        key = item_key(item)
        seen.add(key)
        rec = existing.get(key, {})
        miss = _missing_fields(item)
        rows[key] = {
            "_key": key,
            "shop_name": item.get("shop_name"),
            "name": item.get("name"),
            "volume_cl": _clean(item.get("volume_cl")),
            "_forslag": _volume_hint(item),
            "pack_count": _clean(item.get("pack_count")),
            "brewery": _clean(item.get("brewery")),
            "abv": _clean(item.get("abv")),
            "noter": _sync_brewery_flag(rec.get("noter"), _clean(item.get("brewery")), item.get("shop_name")),
            "_ignorer": rec.get("_ignorer", False),
            "_mangler": "" if rec.get("_ignorer") else ", ".join(miss),
            "_status": "IGNORERET" if rec.get("_ignorer") else ("SMAGEKASSE" if _is_smagekasse(item) else ("MANGLER" if miss else "OK")),
            "_sidst_set": today,
            "_stale": False,
        }
    # Bevar rækker brugeren har rettet, men som ikke kom med i denne scraping
    for key, rec in existing.items():
        if key in seen:
            continue
        rows[key] = {
            "_key": key,
            "shop_name": rec.get("shop_name"),
            "name": rec.get("name"),
            "volume_cl": rec.get("volume_cl"),
            "_forslag": "",
            "pack_count": rec.get("pack_count"),
            "brewery": rec.get("brewery"),
            "abv": rec.get("abv"),
            "noter": _sync_brewery_flag(rec.get("noter"), rec.get("brewery"), rec.get("shop_name")),
            "_ignorer": rec.get("_ignorer", False),
            "_mangler": "",
            "_status": "IGNORERET" if rec.get("_ignorer") else "IKKE SET",
            "_sidst_set": "",
            "_stale": True,
        }

    # Sortér: MANGLER øverst, så OK, så IKKE SET; derefter butik+navn
    order = {"MANGLER": 0, "OK": 1, "IKKE SET": 2, "IGNORERET": 3}
    ordered = sorted(
        rows.values(),
        key=lambda r: (order.get(r["_status"], 9), r.get("shop_name") or "", r.get("name") or ""),
    )

    wb = Workbook()
    _write_data_sheet(wb, ordered, today)
    _write_help_sheet(wb)
    wb.save(path)

    n_missing = sum(1 for r in ordered if r["_status"] == "MANGLER")
    return {"rows": len(ordered), "mangler": n_missing}


def _write_data_sheet(wb, ordered, today):
    ws = wb.active
    ws.title = "Fejlliste"
    edit_keys = {k for (_h, k, e, _w) in COLUMNS if e}

    # Header
    for ci, (head, key, editable, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=head)
        c.fill = EDIT_HEADER if editable else FILL_HEADER
        c.font = DARK_BOLD if editable else WHITE_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Rækker
    for ri, rec in enumerate(ordered, start=2):
        miss = rec["_mangler"]
        miss_set = set()
        if "volumen" in miss: miss_set.add("volume_cl")
        if "bryggeri" in miss: miss_set.add("brewery")
        if "pak-antal" in miss: miss_set.add("pack_count")
        if "abv" in miss: miss_set.add("abv")
        for ci, (head, key, editable, width) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(key))
            c.font = BODY
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    wrap_text=(key in ("name", "noter")))
            if rec["_stale"]:
                c.fill = STALE_ROW
            elif key in miss_set:
                c.fill = MISSING_CELL          # gul = udfyld her
            elif key in edit_keys and editable and rec.get(key) is not None:
                c.fill = OK_CELL

    ws.freeze_panes = "E2"   # lås overskrift + de 4 kontekst-kolonner
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(ordered)+1}"
    ws.row_dimensions[1].height = 30


def _write_help_sheet(wb):
    ws = wb.create_sheet("Læs mig")
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Sådan bruger du fejllisten", True),
        ("", False),
        ("1) Gule celler = mangler. Udfyld dem i kolonnerne med GYLDEN overskrift:", False),
        ("   Volume_cl, Pak_antal, Bryggeri, ABV (ABV er valgfri).", False),
        ("", False),
        ("2) Gem filen som .xlsx (samme navn) og kør scraping igen.", False),
        ("   Dine værdier bruges som facit og vinder over scraperen.", False),
        ("", False),
        ("3) Kun NYE øl med manglende felter får status MANGLER næste gang.", False),
        ("   Det du allerede har rettet, dukker ikke op igen.", False),
        ("", False),
        ("Kolonner med MØRKEGRØN overskrift er automatiske — rør dem ikke:", False),
        ("   Status, Mangler, Butik, Navn, Sidst_set, URL.", False),
        ("   URL er nøglen der binder din rettelse til den rigtige øl.", False),
        ("", False),
        ("Status-værdier:", True),
        ("   MANGLER  = der mangler felter for at den kan matche korrekt.", False),
        ("   OK       = alt nødvendigt er udfyldt.", False),
        ("   IKKE SET = produktet kom ikke med i sidste scraping (måske udgået).", False),
        ("", False),
        ("Tip: Volume_cl = indhold pr. enhed (fx 33). Pak_antal = antal i pakken", False),
        ("(1 = enkelt, 6 = sixpack, 24 = kasse). Pris pr. liter regnes ud fra begge.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, size=12 if bold else 10,
                      color="2E5E4E" if bold else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")