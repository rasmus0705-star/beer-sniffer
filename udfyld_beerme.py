"""
Engangs-udfylder for Beer Me.

Læser fejlliste.xlsx, finder Beer Me-rækker der mangler volumen og/eller ABV,
besøger produktsiden (via beer-me.dk URL'en gemt i nøglekolonnen), parser
'ABV: X%' og 'Flaske/Dåse: X cl' fra beskrivelsen, og skriver værdierne ind.

Skriver til en KOPI (fejlliste_forslag.xlsx) så du kan gennemse resultatet,
før du erstatter originalen. Rækker der ikke kan parses, røres ikke.

Kør:  python udfyld_beerme.py
"""

import re
import html
import time
import shutil
import requests
from urllib.parse import unquote, urlparse, parse_qs
from openpyxl import load_workbook

KILDE = "fejlliste.xlsx"
KOPI = "fejlliste_forslag.xlsx"
PAUSE = 0.6           # sekunder mellem hvert sidekald (høflig mod serveren)
TIMEOUT = 15

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# Produkter der ikke er enkeltøl -> spring over (spar kald)
SKIP_ORD = [
    "glas", "kolbe", "krus", "gave", "gavekort", "abonnement",
    "blandet", "blandede", "valgt af", "beer club", "hver måned",
    "måneder", "smagekasse", "bundle", "pakke", "merchandise",
]


def real_url(vareurl):
    """Træk den rigtige beer-me.dk URL ud af partner-ads klik-linket."""
    if not vareurl:
        return None
    # Hvis det allerede er en beer-me.dk URL, brug den direkte
    if "beer-me.dk" in vareurl and "klikbanner" not in vareurl:
        return vareurl
    q = parse_qs(urlparse(vareurl).query)
    if "htmlurl" in q:
        return unquote(q["htmlurl"][0])
    return None


def parse_side(url):
    """Hent produktside og returnér (volume_cl, abv) — hver kan være None."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
    except Exception as e:
        return None, None, f"fejl: {e}"
    if r.status_code != 200:
        return None, None, f"status {r.status_code}"

    text = re.sub(r"<[^>]+>", " ", r.text)
    text = re.sub(r"\s+", " ", html.unescape(text))

    # ABV: 'ABV: 11,5%' eller 'ABV på 13%'
    abv = None
    m = re.search(r"abv\s*(?:p[åa]\s*|[:\s])\s*(\d+(?:[.,]\d+)?)\s*%", text, re.IGNORECASE)
    if m:
        try:
            abv = float(m.group(1).replace(",", "."))
        except ValueError:
            pass

    # Volumen: 'Flaske: 33 cl' / 'Dåse: 44 cl' / 'Størrelse: 50 cl'
    vol = None
    m = re.search(
        r"(?:flaske|d[åa]se|str[øo]rrelse)[:\s]*(\d+(?:[.,]\d+)?)\s*(cl|ml|l)\b",
        text, re.IGNORECASE,
    )
    if not m:
        # fald tilbage på et hvilket som helst 'X cl' i teksten
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*(cl|ml)\b", text)
    if m:
        val = float(m.group(1).replace(",", "."))
        unit = m.group(2).lower()
        if unit == "l":
            val *= 100
        elif unit == "ml":
            val /= 10
        if 0 < val <= 75:
            vol = val

    return vol, abv, "ok"


def main():
    shutil.copy(KILDE, KOPI)
    wb = load_workbook(KOPI)
    ws = wb["Fejlliste"]
    hdr = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(hdr)}

    c_butik = col["Butik"]
    c_navn = col["Navn"]
    c_vol = col["Volume_cl"]
    c_abv = col["ABV"]
    c_key = col["URL (nøgle — rør ikke)"]

    # Find Beer Me-rækker der mangler volumen og/eller ABV
    opgaver = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, c_butik).value != "Beer Me":
            continue
        navn = str(ws.cell(r, c_navn).value or "")
        mangler_vol = ws.cell(r, c_vol).value in (None, "")
        mangler_abv = ws.cell(r, c_abv).value in (None, "")
        if not (mangler_vol or mangler_abv):
            continue
        if any(s in navn.lower() for s in SKIP_ORD):
            continue
        opgaver.append((r, navn, mangler_vol, mangler_abv))

    print(f"Beer Me-rækker at behandle: {len(opgaver)}")
    print("-" * 50)

    udfyldt_vol = udfyldt_abv = ingen = fejl = 0

    for i, (r, navn, m_vol, m_abv) in enumerate(opgaver, 1):
        key = ws.cell(r, c_key).value
        url = real_url(key)
        if not url:
            fejl += 1
            continue

        vol, abv, status = parse_side(url)

        markeringer = []
        if m_vol and vol is not None:
            ws.cell(r, c_vol).value = vol
            udfyldt_vol += 1
            markeringer.append(f"vol={vol}")
        if m_abv and abv is not None:
            ws.cell(r, c_abv).value = abv
            udfyldt_abv += 1
            markeringer.append(f"abv={abv}")

        if markeringer:
            print(f"  [{i}/{len(opgaver)}] ✓ {navn[:42]:42} {', '.join(markeringer)}")
        else:
            ingen += 1
            print(f"  [{i}/{len(opgaver)}] – {navn[:42]:42} (intet fundet)")

        time.sleep(PAUSE)

    wb.save(KOPI)
    print("-" * 50)
    print(f"Volumen udfyldt: {udfyldt_vol}")
    print(f"ABV udfyldt:     {udfyldt_abv}")
    print(f"Intet fundet:    {ingen}")
    print(f"URL-fejl:        {fejl}")
    print()
    print(f"Gemt til: {KOPI}")
    print("Gennemse den, og hvis den ser rigtig ud:")
    print(f"  erstat {KILDE} med {KOPI} og kør din scraping igen.")


if __name__ == "__main__":
    main()